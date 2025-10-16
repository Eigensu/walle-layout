"""Utilities for parsing XLSX and CSV files for player imports"""
import csv
import io
from typing import List, Dict, Any, BinaryIO
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def normalize_header(header: str) -> str:
    """Normalize header names to lowercase with underscores"""
    return header.strip().lower().replace(" ", "_").replace("-", "_")


def parse_xlsx(file: BinaryIO, header_row: int = 1) -> tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse XLSX file and return headers and rows
    
    Args:
        file: File-like object with Excel content
        header_row: Row number for headers (1-based)
        
    Returns:
        Tuple of (headers, rows) where rows are dicts with normalized keys
    """
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        
        if ws is None:
            raise ValueError("Workbook has no active sheet")
        
        # Read all rows
        all_rows = list(ws.iter_rows(values_only=True))
        
        if len(all_rows) < header_row:
            raise ValueError(f"File has only {len(all_rows)} rows, but header_row is {header_row}")
        
        # Extract headers
        raw_headers = all_rows[header_row - 1]
        headers = [normalize_header(str(h)) if h is not None else f"col_{i}" 
                   for i, h in enumerate(raw_headers)]
        
        # Parse data rows
        rows = []
        for row_idx, row_values in enumerate(all_rows[header_row:], start=header_row + 1):
            # Skip empty rows
            if not any(cell is not None and str(cell).strip() for cell in row_values):
                continue
            
            row_dict: Dict[str, Any] = {"_row_number": row_idx}
            for header, value in zip(headers, row_values):
                # Convert value to appropriate type
                if value is None or (isinstance(value, str) and not value.strip()):
                    row_dict[header] = None
                elif isinstance(value, (int, float)):
                    row_dict[header] = value
                else:
                    row_dict[header] = str(value).strip()
            
            rows.append(row_dict)
        
        wb.close()
        return headers, rows
        
    except InvalidFileException as e:
        raise ValueError(f"Invalid Excel file: {str(e)}")
    except Exception as e:
        raise ValueError(f"Error parsing Excel file: {str(e)}")


def parse_csv(file: BinaryIO, header_row: int = 1) -> tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse CSV file and return headers and rows
    
    Args:
        file: File-like object with CSV content
        header_row: Row number for headers (1-based)
        
    Returns:
        Tuple of (headers, rows) where rows are dicts with normalized keys
    """
    try:
        # Wrap binary file in text wrapper
        text_file = io.TextIOWrapper(file, encoding='utf-8', newline='')
        reader = csv.reader(text_file)
        
        all_rows = list(reader)
        
        if len(all_rows) < header_row:
            raise ValueError(f"File has only {len(all_rows)} rows, but header_row is {header_row}")
        
        # Extract headers
        raw_headers = all_rows[header_row - 1]
        headers = [normalize_header(h) if h else f"col_{i}" 
                   for i, h in enumerate(raw_headers)]
        
        # Parse data rows
        rows = []
        for row_idx, row_values in enumerate(all_rows[header_row:], start=header_row + 1):
            # Skip empty rows
            if not any(cell.strip() for cell in row_values if cell):
                continue
            
            row_dict: Dict[str, Any] = {"_row_number": row_idx}
            for header, value in zip(headers, row_values):
                # Convert value to appropriate type
                if not value or not value.strip():
                    row_dict[header] = None
                else:
                    value = value.strip()
                    # Try to parse as number
                    try:
                        if '.' in value:
                            row_dict[header] = float(value)
                        else:
                            row_dict[header] = int(value)
                    except ValueError:
                        row_dict[header] = value
            
            rows.append(row_dict)
        
        return headers, rows
        
    except UnicodeDecodeError:
        raise ValueError("File is not valid UTF-8. Please save your CSV as UTF-8 encoded.")
    except Exception as e:
        raise ValueError(f"Error parsing CSV file: {str(e)}")


def detect_format(filename: str) -> str:
    """Detect file format from filename"""
    if filename.lower().endswith('.xlsx'):
        return 'xlsx'
    elif filename.lower().endswith('.csv'):
        return 'csv'
    else:
        raise ValueError(f"Unsupported file format: {filename}. Use .xlsx or .csv")
