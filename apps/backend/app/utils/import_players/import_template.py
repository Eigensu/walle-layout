"""Template generation utilities for player import"""
import io
import csv
from typing import BinaryIO, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


# Define standard columns
TEMPLATE_COLUMNS = [
    "name",
    "team", 
    "points",
    "slot_code",
    "slot_name",
    "slot_id",
    "status",
    "image_url",
    "matches",
    "runs",
    "wickets",
]

TEMPLATE_HEADERS = [
    "Name",
    "Team",
    "Points",
    "Slot Code",
    "Slot Name",
    "Slot ID",
    "Status",
    "Image URL",
    "Matches",
    "Runs",
    "Wickets",
]


async def generate_xlsx_template(slot_codes: Optional[list[str]] = None) -> BinaryIO:
    """
    Generate XLSX template with data validations
    
    Args:
        slot_codes: Optional list of slot codes for dropdown
        
    Returns:
        Binary file-like object with XLSX content
    """
    wb = Workbook()
    ws = wb.active
    
    if ws is None:
        raise ValueError("Failed to create worksheet")
    
    ws.title = "Players"
    
    # Style for header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header  # type: ignore[assignment]
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set column widths
    ws.column_dimensions["A"].width = 20  # Name
    ws.column_dimensions["B"].width = 18  # Team
    ws.column_dimensions["C"].width = 10  # Points
    ws.column_dimensions["D"].width = 15  # Slot Code
    ws.column_dimensions["E"].width = 15  # Slot Name
    ws.column_dimensions["F"].width = 25  # Slot ID
    ws.column_dimensions["G"].width = 12  # Status
    ws.column_dimensions["H"].width = 30  # Image URL
    ws.column_dimensions["I"].width = 12  # Matches
    ws.column_dimensions["J"].width = 12  # Runs
    ws.column_dimensions["K"].width = 12  # Wickets
    
    # Data validation for Status (column G)
    status_dv = DataValidation(
        type="list",
        formula1='"Active,Inactive,Injured"',
        allow_blank=True
    )
    status_dv.error = "Please select a valid status"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)
    status_dv.add(f"G2:G5000")
    
    # Data validation for Slot Code if provided (column D)
    if slot_codes:
        slot_formula = '","'.join(slot_codes[:50])  # Limit to 50 for formula length
        slot_dv = DataValidation(
            type="list",
            formula1=f'"{slot_formula}"',
            allow_blank=True
        )
        slot_dv.error = "Please select a valid slot code"
        slot_dv.errorTitle = "Invalid Slot"
        ws.add_data_validation(slot_dv)
        slot_dv.add(f"D2:D5000")
    
    # Add example row
    example_row = [
        "Ankit Shah",
        "DV SPARTANS",
        1000,
        "SLOT 1 (Season)",
        "",
        "",
        "Active",
        "https://example.com/player.jpg",
        25,
        742,
        0,
    ]
    
    for col_idx, value in enumerate(example_row, start=1):
        ws.cell(row=2, column=col_idx).value = value
    
    # Add instructions in a new sheet
    instructions = wb.create_sheet("Instructions")
    instructions.column_dimensions["A"].width = 80
    
    instruction_text = [
        ("Player Import Template Instructions", True),
        ("", False),
        ("Required Fields:", True),
        ("• Name: Player's full name (required)", False),
        ("• Team: Team name (required)", False),
        ("• Points: Player points (required)", False),
        ("", False),
        ("Optional Fields:", True),
        ("• Slot Code/Name/ID: Reference to slot assignment (e.g., 'SLOT 1 (Season)')", False),
        ("• Status: Active, Inactive, or Injured (default: Active)", False),
        ("• Image URL: Player image URL", False),
        ("• Additional stats: Any extra columns (matches, runs, wickets, etc.) will be stored as stats", False),
        ("", False),
        ("Tips:", True),
        ("• Use the dropdown menu for Status", False),
        ("• Delete the example row before importing", False),
        ("• Save file as .xlsx format", False),
        ("• Maximum 5,000 rows per file", False),
    ]
    
    for row_idx, (text, bold) in enumerate(instruction_text, start=1):
        cell = instructions.cell(row=row_idx, column=1)
        cell.value = text
        if bold:
            cell.font = Font(bold=True, size=12)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_csv_template() -> str:
    """Generate CSV template as string"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow(TEMPLATE_COLUMNS)
    
    # Write example row
    writer.writerow([
        "Ankit Shah",
        "DV SPARTANS",
        1000,
        "SLOT 1 (Season)",
        "",
        "",
        "Active",
        "https://example.com/player.jpg",
        25,
        742,
        0,
    ])
    
    return output.getvalue()
